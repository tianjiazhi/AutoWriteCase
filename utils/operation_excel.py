#!/usr/bin/python3
# -*- coding:utf-8 -*-
# author: 田佳志
# email: tianjiazhi@huawei.com
# file: operation_excel.py
# time: 2020/5/5 8:28
# tool: PyCharm 
# desc: 

from openpyxl import *
from openpyxl.styles import PatternFill, colors

class ReaderExcel:
    def __init__(self,file_name):
        self.file_name = file_name
        self.wb = load_workbook(self.file_name)
        sheets = self.wb.sheetnames
        self.sheet = sheets[0]
        self.ws = self.wb[self.sheet]


    def get_cell_value(self, row:int, column:int):
        '''获取单元格中的值'''
        cell_value = self.ws.cell(row=row, column=column).value
        return cell_value


    def get_row_num(self):
        '''获取总行数'''
        rows = self.ws.max_row
        return rows



class WriteExcel:
    def __init__(self,file_name):
        self.file_name = file_name
        self.wb = Workbook()
        self.ws = self.wb.active
        self.wb.save(self.file_name)


    def write_value_to_cell(self, row, colunm, cellvalue):
        """向单元格中写入数据"""
        cell = self.ws.cell(row=row, column=colunm)
        try:
            cell.value = cellvalue
        except:
            red_fill = PatternFill(start_color=colors.RED, end_color=colors.RED, fill_type="solid")
            cell.fill = red_fill
            cell.value = "write_fail"
        finally:
            self.wb.save(self.file_name)



if __name__ == '__main__':
    # reader = ReaderExcel("interface_template.xlsx")
    # value = reader.get_cell_value(2,4)
    # print(value)
    # print(type(value))
    # print(reader.get_row_num())
    writer = WriteExcel('ASD.xlsx')
    writer.write_value_to_cell(1,1,"我是李四")